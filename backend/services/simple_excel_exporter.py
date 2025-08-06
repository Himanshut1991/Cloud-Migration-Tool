#!/usr/bin/env python3
"""Simple Excel export service"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class SimpleExcelExporter:
    """Simple Excel exporter with minimal dependencies"""
    
    def __init__(self, db, models):
        self.db = db
        self.models = models
        self.output_dir = os.path.join(os.path.dirname(__file__), '..', 'exports')
        os.makedirs(self.output_dir, exist_ok=True)
    
    def export_to_excel(self):
        """Export comprehensive migration plan to Excel format"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'migration_plan_{timestamp}.xlsx'
            filepath = os.path.join(self.output_dir, filename)
            
            print(f"📊 Creating comprehensive Excel file: {filename}")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create all sheets with comprehensive data
            self._create_executive_summary_sheet(wb)
            self._create_detailed_servers_sheet(wb)
            self._create_detailed_databases_sheet(wb)
            self._create_detailed_file_shares_sheet(wb)
            self._create_cost_analysis_sheet(wb)
            self._create_migration_timeline_sheet(wb)
            self._create_recommendations_sheet(wb)
            
            wb.save(filepath)
            print(f"✅ Comprehensive Excel file saved: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Excel export error: {str(e)}")
            import traceback
            traceback.print_exc()
            raise Exception(f"Failed to export to Excel: {str(e)}")
    
    def _create_executive_summary_sheet(self, wb):
        """Create comprehensive executive summary sheet"""
        try:
            print("📋 Creating Executive Summary sheet...")
            ws = wb.create_sheet("Executive Summary")
            
            # Styling
            title_font = Font(bold=True, size=16, color="1F4E79")
            header_font = Font(bold=True, size=12, color="1F4E79")
            subheader_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            
            # Title
            ws['A1'] = "Cloud Migration Assessment Report"
            ws['A1'].font = title_font
            ws.merge_cells('A1:E1')
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            ws.merge_cells('A2:E2')
            
            # Get comprehensive data
            Server = self.models.get('Server')
            Database = self.models.get('Database')
            FileShare = self.models.get('FileShare')
            
            servers_count = Server.query.count() if Server else 0
            databases_count = Database.query.count() if Database else 0
            file_shares_count = FileShare.query.count() if FileShare else 0
            
            # Calculate totals
            total_vcpu = 0
            total_ram = 0
            total_disk = 0
            total_db_size = 0
            total_fs_size = 0
            
            if Server:
                servers = Server.query.all()
                for server in servers:
                    total_vcpu += server.vcpu or 0
                    total_ram += server.ram or 0
                    total_disk += server.disk_size or 0
            
            if Database:
                databases = Database.query.all()
                for db in databases:
                    total_db_size += db.size_gb or 0
            
            if FileShare:
                file_shares = FileShare.query.all()
                for fs in file_shares:
                    total_fs_size += fs.total_size_gb or 0
            
            # Infrastructure Overview
            row = 4
            ws[f'A{row}'] = "INFRASTRUCTURE OVERVIEW"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 2
            overview_data = [
                ("Total Servers", servers_count, "Physical and virtual servers"),
                ("Total Databases", databases_count, "Database instances and systems"),
                ("Total File Shares", file_shares_count, "Network storage and file systems"),
                ("Total vCPUs", total_vcpu, "Combined virtual CPU cores"),
                ("Total RAM", f"{total_ram:,.0f} GB", "Combined memory allocation"),
                ("Total Storage", f"{total_disk + total_db_size + total_fs_size:,.0f} GB", "Combined storage requirements")
            ]
            
            for item, value, description in overview_data:
                ws[f'A{row}'] = item
                ws[f'B{row}'] = value
                ws[f'C{row}'] = description
                ws[f'A{row}'].font = subheader_font
                row += 1
            
            # Migration Complexity Assessment
            row += 2
            ws[f'A{row}'] = "MIGRATION COMPLEXITY ASSESSMENT"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 2
            complexity_score = min(10, max(1, (servers_count + databases_count * 1.5 + file_shares_count * 0.5) / 10))
            complexity_level = "Low" if complexity_score < 3 else "Medium" if complexity_score < 7 else "High"
            
            ws[f'A{row}'] = "Complexity Score"
            ws[f'B{row}'] = f"{complexity_score:.1f}/10"
            ws[f'C{row}'] = f"{complexity_level} Complexity"
            row += 1
            
            ws[f'A{row}'] = "Estimated Duration"
            estimated_weeks = 8 + (servers_count * 1.5) + (databases_count * 2.5) + (file_shares_count * 0.5)
            ws[f'B{row}'] = f"{estimated_weeks:.0f} weeks"
            ws[f'C{row}'] = f"Approximately {estimated_weeks/4:.1f} months"
            row += 1
            
            # Cost Estimates
            row += 2
            ws[f'A{row}'] = "PRELIMINARY COST ESTIMATES"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 2
            monthly_compute = servers_count * 200
            monthly_database = databases_count * 300
            monthly_storage = (total_disk + total_db_size + total_fs_size) * 0.1
            monthly_total = monthly_compute + monthly_database + monthly_storage
            annual_total = monthly_total * 12
            
            cost_data = [
                ("Monthly Compute Costs", f"${monthly_compute:,.2f}", f"Based on {servers_count} servers"),
                ("Monthly Database Costs", f"${monthly_database:,.2f}", f"Based on {databases_count} databases"),
                ("Monthly Storage Costs", f"${monthly_storage:,.2f}", f"Based on {total_disk + total_db_size + total_fs_size:,.0f} GB"),
                ("Total Monthly Cost", f"${monthly_total:,.2f}", "Recurring monthly expenses"),
                ("Total Annual Cost", f"${annual_total:,.2f}", "First year estimated cost")
            ]
            
            for item, value, description in cost_data:
                ws[f'A{row}'] = item
                ws[f'B{row}'] = value
                ws[f'C{row}'] = description
                ws[f'A{row}'].font = subheader_font
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            print(f"✅ Executive Summary created with comprehensive data")
            
        except Exception as e:
            print(f"❌ Error creating executive summary: {str(e)}")
            raise
    
    def _create_detailed_servers_sheet(self, wb):
        """Create detailed servers inventory sheet"""
        try:
            print("📋 Creating Detailed Servers sheet...")
            ws = wb.create_sheet("Server Inventory")
            
            # Headers
            headers = [
                "Server ID", "OS Type", "vCPU", "RAM (GB)", "Disk Size (GB)", 
                "Disk Type", "Current Hosting", "Uptime Pattern", "Technology",
                "Application", "Department", "Criticality", "Recommended Instance",
                "Migration Wave", "Notes"
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Get server data
            Server = self.models.get('Server')
            if Server:
                servers = Server.query.all()
                
                for row_idx, server in enumerate(servers, 2):
                    # Basic server info
                    ws.cell(row=row_idx, column=1, value=server.server_id)
                    ws.cell(row=row_idx, column=2, value=server.os_type or "Unknown")
                    ws.cell(row=row_idx, column=3, value=server.vcpu or 2)
                    ws.cell(row=row_idx, column=4, value=server.ram or 4)
                    ws.cell(row=row_idx, column=5, value=server.disk_size or 100)
                    ws.cell(row=row_idx, column=6, value=server.disk_type or "SSD")
                    ws.cell(row=row_idx, column=7, value=server.current_hosting or "On-Premises")
                    ws.cell(row=row_idx, column=8, value=server.uptime_pattern or "24/7")
                    ws.cell(row=row_idx, column=9, value=server.technology or "General")
                    
                    # Additional details
                    ws.cell(row=row_idx, column=10, value=f"App-{server.server_id}")
                    ws.cell(row=row_idx, column=11, value=self._get_department_for_server(server))
                    ws.cell(row=row_idx, column=12, value=self._get_criticality_for_server(server))
                    ws.cell(row=row_idx, column=13, value=self._get_recommended_instance(server))
                    ws.cell(row=row_idx, column=14, value=self._get_migration_wave(server))
                    ws.cell(row=row_idx, column=15, value=self._get_migration_notes(server))
                
                print(f"✅ Added {len(servers)} servers to detailed inventory")
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
        except Exception as e:
            print(f"❌ Error creating servers sheet: {str(e)}")
            raise
    
    def _create_detailed_databases_sheet(self, wb):
        """Create detailed databases inventory sheet"""
        try:
            print("📋 Creating Detailed Databases sheet...")
            ws = wb.create_sheet("Database Inventory")
            
            headers = [
                "Database Name", "Database Type", "Version", "Size (GB)", "HA/DR Required",
                "Backup Frequency", "Performance Tier", "Current Server", "Dependencies",
                "Recommended RDS", "Migration Method", "Downtime Window", "Complexity",
                "Estimated Cost/Month", "Migration Priority"
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            Database = self.models.get('Database')
            if Database:
                databases = Database.query.all()
                
                for row_idx, db in enumerate(databases, 2):
                    ws.cell(row=row_idx, column=1, value=db.db_name)
                    ws.cell(row=row_idx, column=2, value=db.db_type or "SQL Server")
                    ws.cell(row=row_idx, column=3, value=self._get_db_version(db))
                    ws.cell(row=row_idx, column=4, value=db.size_gb or 100)
                    ws.cell(row=row_idx, column=5, value="Yes" if db.ha_dr_required else "No")
                    ws.cell(row=row_idx, column=6, value=db.backup_frequency or "Daily")
                    ws.cell(row=row_idx, column=7, value=db.performance_tier or "Standard")
                    ws.cell(row=row_idx, column=8, value=f"Server-{db.db_name.split('DB')[0] if 'DB' in db.db_name else '001'}")
                    ws.cell(row=row_idx, column=9, value=self._get_db_dependencies(db))
                    ws.cell(row=row_idx, column=10, value=self._get_recommended_rds(db))
                    ws.cell(row=row_idx, column=11, value=self._get_migration_method(db))
                    ws.cell(row=row_idx, column=12, value=self._get_downtime_window(db))
                    ws.cell(row=row_idx, column=13, value=self._get_db_complexity(db))
                    ws.cell(row=row_idx, column=14, value=f"${self._estimate_db_cost(db):,.2f}")
                    ws.cell(row=row_idx, column=15, value=self._get_db_priority(db))
                
                print(f"✅ Added {len(databases)} databases to detailed inventory")
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16
            
        except Exception as e:
            print(f"❌ Error creating databases sheet: {str(e)}")
            raise
    
    def _create_detailed_file_shares_sheet(self, wb):
        """Create detailed file shares inventory sheet"""
        try:
            print("📋 Creating File Shares sheet...")
            ws = wb.create_sheet("File Share Inventory")
            
            headers = [
                "Share Name", "Total Size (GB)", "File Count", "Access Pattern",
                "File Types", "Access Frequency", "Current Location", "User Count",
                "Department", "Compliance Requirements", "Recommended Storage",
                "Migration Method", "Sync Requirements", "Estimated Cost/Month"
            ]
            
            # Style headers  
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="E1952B", end_color="E1952B", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            FileShare = self.models.get('FileShare')
            if FileShare:
                file_shares = FileShare.query.all()
                
                for row_idx, fs in enumerate(file_shares, 2):
                    ws.cell(row=row_idx, column=1, value=fs.share_name)
                    ws.cell(row=row_idx, column=2, value=fs.total_size_gb or 500)
                    ws.cell(row=row_idx, column=3, value=fs.file_count or 10000)
                    ws.cell(row=row_idx, column=4, value=fs.access_pattern or "Mixed")
                    ws.cell(row=row_idx, column=5, value=fs.file_types or "Office Documents, Media")
                    ws.cell(row=row_idx, column=6, value=fs.access_frequency or "Daily")
                    ws.cell(row=row_idx, column=7, value=self._get_share_location(fs))
                    ws.cell(row=row_idx, column=8, value=self._get_user_count(fs))
                    ws.cell(row=row_idx, column=9, value=self._get_fs_department(fs))
                    ws.cell(row=row_idx, column=10, value=self._get_compliance_requirements(fs))
                    ws.cell(row=row_idx, column=11, value=self._get_recommended_storage(fs))
                    ws.cell(row=row_idx, column=12, value=self._get_fs_migration_method(fs))
                    ws.cell(row=row_idx, column=13, value=self._get_sync_requirements(fs))
                    ws.cell(row=row_idx, column=14, value=f"${self._estimate_fs_cost(fs):,.2f}")
                
                print(f"✅ Added {len(file_shares)} file shares to detailed inventory")
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
            
        except Exception as e:
            print(f"❌ Error creating file shares sheet: {str(e)}")
            raise

if __name__ == "__main__":
    # Test the simple exporter
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    
    from models_new import init_models
    from flask import Flask
    from flask_sqlalchemy import SQLAlchemy
    
    app = Flask(__name__)
    basedir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "migration_tool.db")}'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    
    db = SQLAlchemy(app)
    
    with app.app_context():
        models = init_models(db)
        exporter = SimpleExcelExporter(db, models)
        filepath = exporter.export_to_excel()
        print(f"🎉 Test completed! File: {filepath}")
