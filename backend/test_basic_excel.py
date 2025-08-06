#!/usr/bin/env python3
"""Simple Excel export test"""

from openpyxl import Workbook
import os
from datetime import datetime

def test_basic_excel():
    """Test basic Excel functionality"""
    print("🧪 Testing Basic Excel Export...")
    
    try:
        # Create a simple workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some data
        ws['A1'] = "Test"
        ws['B1'] = "Export"
        ws['A2'] = "Date"
        ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Save the file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'test_excel_{timestamp}.xlsx'
        exports_dir = os.path.join(os.path.dirname(__file__), 'exports')
        filepath = os.path.join(exports_dir, filename)
        
        wb.save(filepath)
        
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            print(f"✅ Basic Excel export successful!")
            print(f"📄 File: {filename}")
            print(f"📊 Size: {file_size:,} bytes")
            return True
        else:
            print("❌ Excel file was not created")
            return False
            
    except Exception as e:
        print(f"❌ Basic Excel export failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_basic_excel()
    if success:
        print("\n🔍 openpyxl is working correctly")
        print("   The issue might be in the ExportService implementation")
    else:
        print("\n❌ openpyxl has issues")
        print("   Check if openpyxl is properly installed")
